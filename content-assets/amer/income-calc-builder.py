#!/usr/bin/env python3
# Amer - The Income Plan Calculator (v2: rental-income goal, scenario A/B/C)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

GREEN = "1E3D2F"; GOLD = "C9A84C"; LGOLD = "E8C96B"; CREAM = "F5EDE0"
YELL = "FFF2CC"; REDF = "F4CCCC"; HITF = "F3E4B5"
def F(**k):
    k.setdefault("size", 11)
    return Font(name="Calibri", **k)
HD = F(bold=True, color=GOLD)
INP = F(color="0000FF")  # blue = hardcoded input
CALC = F()
fillH = PatternFill("solid", fgColor=GREEN)
fillY = PatternFill("solid", fgColor=YELL)
thin = Side(style="thin", color="B7C4BB")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Income Plan"

def put(cell, val, font=None, fill=None, fmt=None, align=None, border=False):
    c = ws[cell]; c.value = val
    c.font = font or CALC
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    if align: c.alignment = Alignment(horizontal=align)
    if border: c.border = box
    return c

# ---------- Title ----------
put("A1", "AMER  -  THE INCOME PLAN CALCULATOR", Font(name="Palatino Linotype", size=16, bold=True, color=GREEN))
put("A2", "Goal: D1.5-2.0M net rental income per year by 2036 (D = AED). Edit yellow cells only. Blue = input, black = formula.", F(italic=True, color="666666"))

# ---------- Inputs ----------
put("A4", "INPUTS", HD, fillH)
ws.merge_cells("A4:C4")
inputs = [
    ("AED per GBP", 5.0, "0.0"),
    ("Appreciation % per yr (from 2027)", 0.06, "0.0%"),
    ("Airbnb net yield (% of value)", 0.10, "0.0%"),
    ("Long-let net yield (% of value)", 0.05, "0.0%"),
    ("Flip: avg cash paid-in %", 0.35, "0.0%"),
    ("Flip: selling cost %", 0.02, "0.0%"),
    ("Flip: cycle length (yrs)", 3, "0"),
    ("Target net income (D per yr)", 2000000, "#,##0"),
    ("Scenario (A / B / C)", "A", None),
]
r = 5
for label, val, fmt in inputs:
    put(f"A{r}", label)
    put(f"B{r}", val, INP, fillY, fmt, border=True)
    r += 1
put("A14", "Flip ROE per cycle (calculated)")
put("B14", "=((1+B6)^B11-1-B10)/B9", CALC, None, "0.0%", border=True)
put("A15", "P3 mortgage service (D per yr)")
put("B15", 160000, INP, fillY, "#,##0", border=True)
put("A16", "P3 mortgage service last year")
put("B16", 2036, INP, fillY, "0", border=True)
put("A17", "Use P3 mortgage (Y / N)")
put("B17", "Y", INP, fillY, None, "center", border=True)
put("A18", "Y: 60% handover balance (D2.4M) mortgaged, service above, principal cleared in the year at B16. N: 60% paid in cash at 2030 - then P4 ~2033, P5 ~2036 keeps the pot green. Terra/Athlon instalments (~D2.0M to 2029, est. 20% paid) are committed spend outside this model.", F(italic=True, size=8, color="666666"))

dv = DataValidation(type="list", formula1='"A,B,C"', allow_blank=False)
ws.add_data_validation(dv); dv.add("B13")
dv3 = DataValidation(type="list", formula1='"Y,N"', allow_blank=False)
ws.add_data_validation(dv3); dv3.add("B17")

# ---------- Cash-in scenarios ----------
put("F4", "CASH IN FROM AMER (GBP '000 per year)", HD, fillH)
ws.merge_cells("F4:J4")
for col, h in zip("FGHIJ", ["Year", "A", "B", "C", "Selected (D)"]):
    put(f"{col}5", h, F(bold=True), None, None, "center", border=True)
scenA = [500,100,100,100,0,0,0,0,0,0]
scenB = [300,200,200,200,0,0,0,0,0,0]
scenC = [100]*10
for i in range(10):
    rr = 6 + i
    put(f"F{rr}", 2026+i, CALC, None, "0", "center", border=True)
    put(f"G{rr}", scenA[i], INP, fillY, "#,##0", border=True)
    put(f"H{rr}", scenB[i], INP, fillY, "#,##0", border=True)
    put(f"I{rr}", scenC[i], INP, fillY, "#,##0", border=True)
    put(f"J{rr}", f"=INDEX(G{rr}:I{rr},MATCH($B$13,$G$5:$I$5,0))*$B$5*1000", CALC, None, "#,##0", border=True)
put("F16", "Total", F(bold=True), border=True)
for col in "GHI":
    put(f"{col}16", f"=SUM({col}6:{col}15)", F(bold=True), None, "#,##0", border=True)
put("J16", "=SUM(J6:J15)", F(bold=True), None, "#,##0", border=True)
put("F17", "Option C underfunds this plan (pot goes red below). To see an honest C: shrink P3 to ~D3M and push flip buy years later.", F(italic=True, size=9, color="666666"))

# ---------- Positions ----------
put("A19", "THE POSITIONS  (edit prices and years - the plan recalculates)", HD, fillH)
ws.merge_cells("A19:O19")
heads = ["Position","Role","Price D","Buy yr","Income from","Exit yr (flips)",
         "Value at start/exit D","Net income at start D/yr","Cash back at exit D",
         "New cash to deploy D","Deploy until (yr)","Value 2036 D","Income 2036 D","Svc chg D/yr","Notes"]
for j, h in enumerate(heads):
    put(f"{get_column_letter(1+j)}20", h, F(bold=True, size=9), None, None, "center", border=True)

# rows 21-29
pos = [
    # name, role, price, buy, incomeFrom, exitYr, J-formula(deploy), K-formula(until), note
    ("Terra Gardens 1-bed (owned)", "Airbnb", 1640000, 2025, 2030, None, "0", "=D{r}+1",
     "Owned. Airbnb from handover Q4 2029. Remaining instalments already committed - not new cash."),
    ("Athlon 1-bed (owned - SELL)", "Flip", 1823000, 2025, None, 2029, "0", "=F{r}",
     "Owned. Sell at handover. Cash back = sale - 40% balance owed - 2% sale cost. Decision gate 2029."),
    ("P3 keeper - Hudayriyat D4M", "Airbnb", 4000000, 2026, 2030, None, "=0.42*C{r}", "=E{r}",
     "Deploy = 40% pre-handover + 2% ADM. 60% handover balance mortgaged (inputs B15/B16)."),
    ("F1 flip - AD off-plan", "Flip", 2000000, 2026, None, 2029, "=($B$9+0.02)*C{r}", "=F{r}",
     "Deploy = paid-in % + 2% ADM entry. Cash back = paid-in + full-value appreciation - sale cost."),
    ("F2 flip - AD off-plan", "Flip", 2000000, 2027, None, 2030, "=($B$9+0.02)*C{r}", "=F{r}",
     "Second flip, staggered exit."),
    ("F3 flip - off-plan", "Flip", 2000000, 2030, None, 2033, "=($B$9+0.02)*C{r}", "=F{r}",
     "Last flip cycle - hard stop on new flips after this."),
    ("P4 keeper - ready unit", "Airbnb", 4000000, 2030, 2031, None, "=1.04*C{r}", "=D{r}+1",
     "Ready unit bought with Athlon money + pot. Full price + 4% DLD. Income within a year."),
    ("P5 keeper - ready unit", "Long-let", 4000000, 2034, 2034, None, "=1.04*C{r}", "=D{r}+1",
     "Bought early 2034 with the F3 exit + accumulated rents - service charges make 2033 too tight to buy in-year."),
    ("P6 keeper - ready unit", "Long-let", 2000000, 2035, 2036, None, "=1.04*C{r}", "=D{r}+1",
     "Final rung, from surplus rents."),
]
svcmap = [15000,0,10000,0,0,0,15000,10000,15000]
for i, (name, role, price, buy, incf, exy, jf, kf, note) in enumerate(pos):
    r = 21 + i
    put(f"A{r}", name, F(size=10), border=True)
    put(f"B{r}", role, INP, fillY, None, "center", border=True)
    put(f"C{r}", price, INP, fillY, "#,##0", border=True)
    put(f"D{r}", buy, INP, fillY, "0", "center", border=True)
    put(f"E{r}", incf if incf else "", INP, fillY, "0", "center", border=True)
    put(f"F{r}", exy if exy else "", INP, fillY, "0", "center", border=True)
    put(f"G{r}", f'=IF($F{r}<>"",$C{r}*(1+$B$6)^($F{r}-MAX($D{r},2026)),IF($E{r}<>"",$C{r}*(1+$B$6)^($E{r}-MAX($D{r},2026)),0))', CALC, None, "#,##0", border=True)
    put(f"H{r}", f'=IF($E{r}="",0,IF($B{r}="Airbnb",$B$7,IF($B{r}="Long-let",$B$8,0))*$G{r}-$N{r})', CALC, None, "#,##0", border=True)
    if name.startswith("Athlon"):
        put(f"I{r}", f'=IF($F{r}="",0,$G{r}-0.4*$C{r}-$B$10*$G{r})', CALC, None, "#,##0", border=True)
    else:
        put(f"I{r}", f'=IF($F{r}="",0,$C{r}*$B$9+$G{r}-$C{r}-$B$10*$G{r})', CALC, None, "#,##0", border=True)
    put(f"J{r}", jf.format(r=r) if jf.startswith("=") else 0, CALC, None, "#,##0", border=True)
    put(f"K{r}", kf.format(r=r), CALC, None, "0", "center", border=True)
    put(f"L{r}", f'=IF($F{r}<>"",0,$C{r}*(1+$B$6)^(2036-MAX($D{r},2026)))', CALC, None, "#,##0", border=True)
    put(f"M{r}", f'=IF($E{r}="",0,IF($B{r}="Airbnb",$B$7,IF($B{r}="Long-let",$B$8,0))*$L{r}-$N{r})', CALC, None, "#,##0", border=True)
    put(f"N{r}", svcmap[i], INP, fillY, "#,##0", border=True)
    put(f"O{r}", note, F(size=8, color="666666"), border=True)
put("A30", "TOTALS", F(bold=True), border=True)
for col in ("I","J","L","M"):
    put(f"{col}30", f"=SUM({col}21:{col}29)", F(bold=True), None, "#,##0", border=True)

dv2 = DataValidation(type="list", formula1='"Airbnb,Long-let,Flip"', allow_blank=False)
ws.add_data_validation(dv2); dv2.add("B21:B29")

# ---------- Timeline ----------
put("A33", "THE TIMELINE  (all D)", HD, fillH)
ws.merge_cells("A33:N33")
put("A34", "Year", F(bold=True), border=True)
for j in range(13):
    put(f"{get_column_letter(2+j)}34", 2026+j, F(bold=True), None, "0", "center", border=True)
# income per position rows 35-43 reference positions rows 21-29
for i in range(9):
    tr = 35 + i; pr = 21 + i
    put(f"A{tr}", f"=A{pr}", F(size=9))
    for j in range(13):
        col = get_column_letter(2+j)
        put(f"{col}{tr}", f'=IF(OR($E{pr}="",{col}$34<$E{pr}),0,IF($B{pr}="Airbnb",$B$7,IF($B{pr}="Long-let",$B$8,0))*$C{pr}*(1+$B$6)^({col}$34-MAX($D{pr},2026))-$N{pr})', F(size=9), None, "#,##0")
labels = [
    (44, "Total net rent", "=SUM({c}35:{c}43)", "#,##0", True),
    (45, "Less: P3 mortgage service", '=IF(AND($B$17="Y",{c}$34>=2030,{c}$34<=$B$16),-$B$15,0)', "#,##0;(#,##0)", False),
    (46, "TOTAL NET INCOME (D)", "={c}44+{c}45", "#,##0", True),
    (47, "   in GBP '000", "={c}46/$B$5/1000", "#,##0", False),
    (48, "Milestone", '=IF({c}46>=$B$12,"TARGET",IF({c}46>=1500000,"D1.5M",""))', None, False),
    (49, "Cash in from Amer", '=IF(COUNTIF($F$6:$F$15,{c}$34)=0,0,INDEX($J$6:$J$15,MATCH({c}$34,$F$6:$F$15,0)))', "#,##0", False),
    (50, "Flip cash back", "=SUMIFS($I$21:$I$29,$F$21:$F$29,{c}$34)", "#,##0", False),
    (51, "Cash deployed", "=SUMPRODUCT(($D$21:$D$29<={c}$34)*($K$21:$K$29>{c}$34)*($J$21:$J$29)/($K$21:$K$29-$D$21:$D$29))", "#,##0", False),
    (52, "P3 handover 60% / principal", '=IF($B$17="Y",IF({c}$34=$B$16,0.6*$C$23,0),IF({c}$34=2030,0.6*$C$23,0))', "#,##0", False),
    (54, "Funding check", '=IF({c}53<0,"UNDERFUNDED","ok")', None, False),
]
for rr, lab, f_t, fmt, bold in labels:
    put(f"A{rr}", lab, F(bold=bold))
    for j in range(13):
        col = get_column_letter(2+j)
        put(f"{col}{rr}", f_t.format(c=col), F(bold=bold, size=10), None, fmt, "center" if fmt is None else None)
# pot row 53 (cumulative), cumulative invested row 55
put("A53", "Cash pot (cumulative)", F(bold=True))
put("B53", "=B46+B49+B50-B51-B52", F(bold=True, size=10), None, "#,##0")
put("A55", "Cumulative cash deployed", F())
put("B55", "=B51+B52", F(size=10), None, "#,##0")
for j in range(1, 13):
    col = get_column_letter(2+j); prev = get_column_letter(1+j)
    put(f"{col}53", f"={prev}53+{col}46+{col}49+{col}50-{col}51-{col}52", F(bold=True, size=10), None, "#,##0")
    put(f"{col}55", f"={prev}55+{col}51+{col}52", F(size=10), None, "#,##0")

# ---------- Read-out ----------
put("A58", "READ-OUT", HD, fillH)
ws.merge_cells("A58:C58")
ro = [
    ("Net income 2036 (D)", "=INDEX($B$46:$N$46,MATCH(2036,$B$34:$N$34,0))", "#,##0"),
    ("Year D1.5M crossed", '=IF(COUNTIF($B$46:$N$46,"<"&1500000)>=13,"after 2038",2026+COUNTIF($B$46:$N$46,"<"&1500000))', "0"),
    ("Year target crossed", '=IF(COUNTIF($B$46:$N$46,"<"&$B$12)>=13,"after 2038",2026+COUNTIF($B$46:$N$46,"<"&$B$12))', "0"),
    ("Portfolio value 2036 (D)", "=L30", "#,##0"),
    ("Total new cash in (GBP '000)", "=SUM($J$6:$J$15)/$B$5/1000", "#,##0"),
    ("Cash pot end 2036 (D)", "=INDEX($B$53:$N$53,MATCH(2036,$B$34:$N$34,0))", "#,##0"),
    ("Cash pot end 2038 (D)", "=N53", "#,##0"),
]
r = 59
for lab, f_t, fmt in ro:
    put(f"A{r}", lab, F(bold=True))
    put(f"B{r}", f_t, F(bold=True, color="7A6018"), fillY, fmt, border=True)
    r += 1

# ---------- Legend ----------
put("A68", "HOW TO USE", HD, fillH)
ws.merge_cells("A68:C68")
legend = [
    "1. Pick scenario A / B / C in the yellow box (B13) - it changes the cash-in column only.",
    "2. Yellow cells are yours to edit: yields, appreciation, prices, buy/exit years, cash per year.",
    "3. The Funding check row goes UNDERFUNDED (red) when the pot runs dry - push buy years later or add cash.",
    "4. Flip cash back assumes appreciation on FULL unit value while only the paid-in % is your cash - that is the engine.",
    "5. Mortgage toggle (B17): N pays P3 60% in cash at 2030 - then P4 ~2033 and P5 ~2036 (or P4 at D3M in 2032) keeps the pot green.",
    "6. Yields are before service charges - the Svc chg column deducts D10K/yr (townhouse-villa) or D15-20K/yr (apartments) per unit.",
    "7. All figures indicative. Appreciation and yields are assumptions from Calum's 4-7 Aug 2026 notes, not promises.",
]
for i, t in enumerate(legend):
    put(f"A{69+i}", t, F(size=9, color="666666"))

# ---------- Conditional formatting ----------
ws.conditional_formatting.add("B46:N46",
    FormulaRule(formula=['B46>=$B$12'], fill=PatternFill("solid", fgColor=HITF)))
ws.conditional_formatting.add("B53:N54",
    FormulaRule(formula=['B$53<0'], fill=PatternFill("solid", fgColor=REDF)))

# ---------- Widths ----------
ws.column_dimensions["A"].width = 30
for col in "BCDEFGHIJKLM":
    ws.column_dimensions[col].width = 12
ws.column_dimensions["N"].width = 11
ws.column_dimensions["O"].width = 40
ws.freeze_panes = "B5"

wb.save("/home/claude/work/amer/Amer-Income-Plan-Calculator.xlsx")
print("saved")
